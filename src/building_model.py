import os
from copy import deepcopy
from datetime import datetime

import numpy as np
import openpyxl
import pandas as pd
import xlwt
import matplotlib.pyplot as plt
import logging as log

from odym.modules import ODYM_Classes as msc
from odym.modules import ODYM_Functions as msf


def main():
    data_path = add_docs_path()

    results_path = add_results_path()

    mylog = initialize_logging()

    model_configsheet, script_config = read_parameters_from_config(data_path, mylog)

    it_aspects, it_description, it_dimension, it_index_letter, model_classification, pl_index_layer, pl_index_match, \
        pl_index_structure, pl_names, pl_version, pr_l_name, pr_l_number, master_classification, script_config = \
        read_parameters_from_classification(data_path, model_configsheet, mylog, script_config)

    model_time_end, model_time_start = define_model_time(model_classification, mylog)

    index_table, index_table_classification_names = define_index_table(it_aspects, it_description, it_dimension,
                                                                       it_index_letter, model_classification, mylog)

    parameter_dict = read_data_and_parameters(data_path, index_table, index_table_classification_names,
                                              master_classification, mylog, pl_index_layer, pl_index_match,
                                              pl_index_structure, pl_names, pl_version, script_config)

    building_mfa_system = define_mfa_system(index_table, model_time_end, model_time_start, parameter_dict, mylog)

    add_processes_mfa(building_mfa_system, pr_l_name, pr_l_number, mylog)

    add_flows_mfa(building_mfa_system, mylog)

    add_stocks_mfa(building_mfa_system, mylog)

    building_mfa_system.Initialize_FlowValues()

    building_mfa_system.Initialize_StockValues()

    building_mfa_system.Consistency_Check()

    reference_calculation(building_mfa_system, results_path, mylog)

    ce_action_calculation(building_mfa_system, results_path, mylog)

    ce_bundle_calculation(building_mfa_system, results_path, mylog)


def add_docs_path():
    data_path = os.path.join(os.getcwd(), '../..', 'buildings_pro_stock_EU', 'docs')
    return data_path


def add_results_path():
    folder_path = os.path.join(os.getcwd(), '../..', 'buildings_pro_stock_EU/results/')
    today = datetime.today().strftime('%Y-%m-%d')
    results_path = folder_path + 'results_' + today + '/'
    os.mkdir(results_path)
    return results_path


def initialize_logging():
    log_verbosity = eval("log.DEBUG")
    log_filename = 'LogFileTest.md'
    [mylog, console_log, file_log] = msf.function_logger(log_filename, os.getcwd(),
                                                         log_verbosity, log_verbosity)
    return mylog


def read_parameters_from_config(data_path, mylog):
    mylog.info('Read and parse config table, including the model index table, from model config sheet.')
    project_specs_name_con_file = 'config_stock_model.xlsx'
    model_configfile = openpyxl.load_workbook(os.path.join(data_path, project_specs_name_con_file), data_only=True)
    script_config = {'Model Setting': model_configfile['Config'].cell(4, 4).value}
    model_configsheet = model_configfile[script_config['Model Setting']]
    name_scenario = model_configsheet.cell(4, 4).value
    print(name_scenario)
    script_config = msf.ParseModelControl(model_configsheet, script_config)
    print(script_config)
    return model_configsheet, script_config


def read_parameters_from_classification(data_path, model_configsheet, mylog, script_config):
    mylog.info('Define model classifications and select items for model classifications according to information '
               'provided by classification file')
    project_specs_name_classfile = 'classifications_stock_model.xlsx'
    classfile = openpyxl.load_workbook(os.path.join(data_path, project_specs_name_classfile), data_only=True)
    classsheet = classfile['MAIN_Table']
    master_classification = msf.ParseClassificationFile_Main(classsheet, mylog)
    print('Read index table from model config sheet.')
    it_aspects, it_description, it_dimension, it_classification, it_selector, it_index_letter, pl_names, \
        pl_description, pl_version, pl_index_structure, pl_index_match, pl_index_layer, pr_l_number, pr_l_name, \
        pr_l_comment, pr_l_type, script_config = msf.ParseConfigFile(model_configsheet, script_config, mylog)
    print('Define model classifications and select items for model classifications according to '
          'information provided by config file.')
    model_classification = {}
    for m in range(0, len(it_aspects)):
        model_classification[it_aspects[m]] = deepcopy(master_classification[it_classification[m]])
        eval_string = msf.EvalItemSelectString(it_selector[m], len(model_classification[it_aspects[m]].Items))
        if eval_string.find(':') > -1:
            range_start = int(eval_string[0:eval_string.find(':')])
            range_stop = int(eval_string[eval_string.find(':') + 1::])
            model_classification[it_aspects[m]].Items = \
                model_classification[it_aspects[m]].Items[range_start:range_stop]
        elif eval_string.find('[') > -1:
            model_classification[it_aspects[m]].Items = \
                [model_classification[it_aspects[m]].Items[i] for i in eval(eval_string)]
        elif eval_string == 'all':
            pass
        else:
            mylog.info('ITEM SELECT ERROR for aspect ' + it_aspects[m] + ' were found in datafile.</br>')
            break
    return it_aspects, it_description, it_dimension, it_index_letter, model_classification, pl_index_layer, \
        pl_index_match, pl_index_structure, pl_names, pl_version, pr_l_name, pr_l_number, master_classification, \
        script_config


def define_model_time(model_classification, mylog):
    mylog.info('Define model time')
    model_time_start = int(min(model_classification['Time'].Items))
    model_time_end = int(max(model_classification['Time'].Items))
    return model_time_end, model_time_start


def define_index_table(it_aspects, it_description, it_dimension, it_index_letter, model_classification, mylog):
    mylog.info('Define index table dataframe')
    print('Define index table dataframe.')
    index_table = pd.DataFrame({'Aspect': it_aspects,
                                'Description': it_description,
                                'Dimension': it_dimension,
                                'Classification': [model_classification[Aspect] for Aspect in it_aspects],
                                'IndexLetter': it_index_letter})
    index_table.set_index('Aspect', inplace=True)
    index_table['IndexSize'] = \
        pd.Series([len(index_table.Classification[i].Items) for i in range(0, len(index_table.IndexLetter))],
                  index=index_table.index)
    index_table_classification_names = [index_table.Classification[i].Name for i in
                                        range(0, len(index_table.IndexLetter))]
    return index_table, index_table_classification_names


def read_data_and_parameters(data_path, index_table, index_table_classification_names, master_classification,
                             mylog, pl_index_layer, pl_index_match, pl_index_structure, pl_names, pl_version,
                             script_config):
    print('Read model data and parameters.')
    parameter_dict = {}
    for mo in range(0, len(pl_names)):
        par_path = os.path.join(data_path, pl_version[mo])
        print('Reading parameter ' + pl_names[mo])
        mylog.info('Reading parameter' + pl_names[mo])
        meta_data, values = msf.ReadParameterXLSX(par_path, pl_names[mo], pl_index_structure[mo],
                                                  pl_index_match[mo], pl_index_layer[mo],
                                                  master_classification, index_table,
                                                  index_table_classification_names, script_config, mylog, False)
        parameter_dict[pl_names[mo]] = msc.Parameter(Name=meta_data['Dataset_Name'],
                                                     ID=meta_data['Dataset_ID'],
                                                     UUID=meta_data['Dataset_UUID'],
                                                     P_Res=None,
                                                     MetaData=meta_data,
                                                     Indices=pl_index_structure[mo],
                                                     Values=values,
                                                     Uncert=None,
                                                     Unit=meta_data['Dataset_Unit'])
    mylog.info('Reading of parameters finished')
    return parameter_dict


def define_mfa_system(index_table, model_time_end, model_time_start, parameter_dict, mylog):
    mylog.info('Define MFA system and processes')
    print('Define MFA system and processes.')
    building_mfa_system = msc.MFAsystem(Name='Building_Model',
                                        Geogr_Scope='EU',
                                        Unit='kg',
                                        ProcessList=[],
                                        FlowDict={},
                                        StockDict={},
                                        ParameterDict=parameter_dict,
                                        Time_Start=model_time_start,
                                        Time_End=model_time_end,
                                        IndexTable=index_table,
                                        Elements=index_table.loc['Element'].Classification.Items,
                                        Graphical=None)
    building_mfa_system.IndexTableCheck()
    return building_mfa_system


def add_processes_mfa(building_mfa_system, pr_l_name, pr_l_number, mylog):
    mylog.info('Add processes to MFA system')
    print('Add processes to MFA system')
    for m in range(0, len(pr_l_number)):
        building_mfa_system.ProcessList.append(msc.Process(Name=pr_l_name[m], ID=pr_l_number[m]))


def add_flows_mfa(building_mfa_system, mylog):
    mylog.info('Add flows to MFA system')
    print('Add flows to MFA system')
    building_mfa_system.FlowDict['Construction of buildings'] = \
        msc.Flow(Name='Construction of Buildings', P_Start=0, P_End=1, Indices='r,b,a,t', Values=None)
    building_mfa_system.FlowDict['Demolition of buildings'] = \
        msc.Flow(Name='Demolition of Buildings', P_Start=1, P_End=0, Indices='r,b,a,t', Values=None)
    building_mfa_system.FlowDict['Steel production'] = \
        msc.Flow(Name='Steel production', P_Start=2, P_End=3, Indices='r,s,t', Values=None)
    building_mfa_system.FlowDict['Steel inflow'] = \
        msc.Flow(Name='Steel inflow', P_Start=3, P_End=4, Indices='r,f,t', Values=None)
    building_mfa_system.FlowDict['Steel inflow, building types'] = \
        msc.Flow(Name='Steel inflow, building types', P_Start=3, P_End=4, Indices='r,f,t', Values=None)
    building_mfa_system.FlowDict['Steel outflow'] = \
        msc.Flow(Name='Steel outflow', P_Start=4, P_End=5, Indices='r,f,t', Values=None)
    building_mfa_system.FlowDict['Scrap recycling'] = \
        msc.Flow(Name='Scrap recycling', P_Start=5, P_End=2, Indices='s,t,r', Values=None)
    building_mfa_system.FlowDict['Scrap other use'] = \
        msc.Flow(Name='Scrap other use', P_Start=5, P_End=2, Indices='r,t', Values=None)
    building_mfa_system.FlowDict['Reuse of steel element'] = \
        msc.Flow(Name='Reuse of steel element', P_Start=4, P_End=4, Indices='r,f,t', Values=None)
    building_mfa_system.FlowDict['Reuse of steel'] = \
        msc.Flow(Name='Reuse of steel', P_Start=4, P_End=4, Indices='r,f,t', Values=None)
    building_mfa_system.FlowDict['Clinker production'] = \
        msc.Flow(Name='Clinker production', P_Start=6, P_End=7, Indices='r,l,t', Values=None)
    building_mfa_system.FlowDict['Cement production'] = \
        msc.Flow(Name='Cement production', P_Start=7, P_End=8, Indices='r,m,t', Values=None)
    building_mfa_system.FlowDict['Concrete inflow'] = \
        msc.Flow(Name='Concrete inflow', P_Start=8, P_End=9, Indices='r,o,t', Values=None)
    building_mfa_system.FlowDict['Concrete inflow, building types'] = \
        msc.Flow(Name='Concrete inflow, building types', P_Start=8, P_End=9, Indices='r,o,t', Values=None)
    building_mfa_system.FlowDict['Concrete outflow'] = \
        msc.Flow(Name='Concrete outflow', P_Start=9, P_End=10, Indices='r,o,t', Values=None)
    building_mfa_system.FlowDict['Concrete reuse'] = \
        msc.Flow(Name='Concrete reuse', P_Start=10, P_End=8, Indices='r,o,t', Values=None)
    building_mfa_system.FlowDict['Concrete landfill'] = \
        msc.Flow(Name='Concrete landfill', P_Start=10, P_End=0, Indices='r,t', Values=None)
    building_mfa_system.FlowDict['Reuse of concrete element'] = \
        msc.Flow(Name='Reuse of concrete elements', P_Start=9, P_End=9, Indices='r,o,t', Values=None)
    building_mfa_system.FlowDict['Cement recycling'] = \
        msc.Flow(Name='Cement recycling', P_Start=10, P_End=7, Indices='r,m,t', Values=None)


def add_stocks_mfa(building_mfa_system, mylog):
    mylog.info('Add stocks to MFA system')
    print('Add stocks to MFA system')
    building_mfa_system.StockDict['Building stock'] = \
        msc.Stock(Name='Building stock', P_Res=1, Type=0, Indices='r,b,a,t', Values=None)
    building_mfa_system.StockDict['Building stock change'] = \
        msc.Stock(Name='Building stock change', P_Res=1, Type=1, Indices='r,b,a,t', Values=None)
    building_mfa_system.StockDict['Steel stock in buildings'] = \
        msc.Stock(Name='Steel stock in buildings', P_Res=4, Type=0, Indices='r,f,t', Values=None)
    building_mfa_system.StockDict['Steel stock change in buildings'] = \
        msc.Stock(Name='Steel stock change in buildings', P_Res=4, Type=1, Indices='r,b,a,t', Values=None)
    building_mfa_system.StockDict['Concrete stock in buildings'] = \
        msc.Stock(Name='Concrete stock in buildings', P_Res=9, Type=0, Indices='r,o,t', Values=None)
    building_mfa_system.StockDict['Concrete stock change in buildings'] = \
        msc.Stock(Name='Concrete stock change in buildings', P_Res=9, Type=1, Indices='r,o,t', Values=None)


def reference_calculation(building_mfa_system, results_path, mylog):
    solve_mfa_reference(building_mfa_system, mylog)
    write_results_excel_reference(building_mfa_system, results_path, mylog)


def solve_mfa_reference(building_mfa_system, mylog):
    mylog.info('Solve MFA for reference case')
    print('Solve MFA for reference case')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.FlowDict['Steel inflow, building types'].Values = \
        np.einsum('rbaf, rbat->bt',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow, building types'].Values = \
        np.einsum('rbao, rbat ->bt',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_reference(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for reference case')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for reference case')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_in_ty', building_mfa_system.FlowDict['Steel inflow, building types'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['BuildingType'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_in_ty', building_mfa_system.FlowDict['Concrete inflow, building types'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['BuildingType'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/reference.xls')


def ce_action_calculation(building_mfa_system, results_path, mylog):
    calc_timber_construction(building_mfa_system, results_path, mylog)
    calc_reduced_space(building_mfa_system, results_path, mylog)
    calc_reduced_overspec(building_mfa_system, results_path, mylog)
    calc_cult_herit(building_mfa_system, results_path, mylog)
    calc_renovation(building_mfa_system, results_path, mylog)
    calc_reuse_elements(building_mfa_system, results_path, mylog)
    calc_reuse_steel(building_mfa_system, results_path, mylog)
    calc_rec_cement(building_mfa_system, results_path, mylog)


def calc_timber_construction(building_mfa_system, results_path, mylog):
    solve_mfa_timber_construction(building_mfa_system, mylog)
    write_results_excel_timber_construction(building_mfa_system, results_path, mylog)


def solve_mfa_timber_construction(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to timber construction')
    print('Calculate reduced material demand due to timber construction')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_timber'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_timber'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_timber'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_timber'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_timber'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_timber'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_timber_construction(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE case (timber construction)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE case (timber construction)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)

    results_file.save(results_path + '/CE_timberconstruction.xls')


def calc_reduced_space(building_mfa_system, results_path, mylog):
    solve_mfa_reduced_space(building_mfa_system, mylog)
    write_results_excel_reduced_space(building_mfa_system, results_path, mylog)


def solve_mfa_reduced_space(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to reduced floor space demand')
    print('Calculate reduced material demand due to reduced floor space demand')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow_reduced'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock_reduced'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow_reduced'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_reduced_space(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE case (reduced floor space)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE case (reduced floor space)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)

    results_file.save(results_path + '/CE_reducedspace.xls')


def calc_reduced_overspec(building_mfa_system, results_path, mylog):
    solve_mfa_reduced_overspec(building_mfa_system, mylog)
    write_results_excel_reduced_overspec(building_mfa_system, results_path, mylog)


def solve_mfa_reduced_overspec(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to reduced over-specification')
    print('Calculate reduced material demand due to reduced over-specification')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_overspec'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_overspec'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_overspec'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_overspec'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_overspec'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_overspec'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_reduced_overspec(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE case (reduced over-specification)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE case (reduced over-specification)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_reducedoverspec.xls')


def calc_cult_herit(building_mfa_system, results_path, mylog):
    solve_mfa_cult_herit(building_mfa_system, mylog)
    write_results_excel_cult_herit(building_mfa_system, results_path, mylog)


def solve_mfa_cult_herit(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to protection of cultural heritage buildings')
    print('Calculate reduced material demand due to protection of cultural heritage buildings')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow_cult'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock_cult'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow_cult'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_cult_herit(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE case (cultural heritage)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE case (cultural heritage)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_culturalheritage.xls')


def calc_renovation(building_mfa_system, results_path, mylog):
    solve_mfa_renovation(building_mfa_system, mylog)
    write_results_excel_renovation(building_mfa_system, results_path, mylog)


def solve_mfa_renovation(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to renovation')
    print('Calculate reduced material demand due to renovation')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow_renov'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock_renov'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow_renov'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_renovation(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE case (renovation)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE case (renovation)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_renovation.xls')


def calc_reuse_elements(building_mfa_system, results_path, mylog):
    solve_mfa_reuse_elements(building_mfa_system, mylog)
    write_results_excel_reuse_elements(building_mfa_system, results_path, mylog)


def solve_mfa_reuse_elements(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to reuse of building elements')
    print('Calculate reduced material demand due to reuse of building elements')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Reuse of steel element'].Values = \
        np.einsum('rft, rft -> rft',
                  building_mfa_system.ParameterDict['par_steel_element_reuse'].Values,
                  building_mfa_system.FlowDict['Steel outflow'].Values)
    building_mfa_system.FlowDict['Steel inflow'].Values = building_mfa_system.FlowDict['Steel inflow'].Values + \
        building_mfa_system.FlowDict['Reuse of steel element'].Values
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel element'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Reuse of concrete element'].Values = \
        np.einsum('rot, rot -> rot',
                  building_mfa_system.ParameterDict['par_concrete_element_reuse'].Values,
                  building_mfa_system.FlowDict['Concrete outflow'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        building_mfa_system.FlowDict['Concrete inflow'].Values + \
        building_mfa_system.FlowDict['Reuse of concrete element'].Values
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Reuse of concrete element'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_reuse_elements(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE case (reuse of elements)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE case (reuse of elements)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of steel element'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel element'].Values)
    msf.ExcelSheetFill(results_file, 'S_el_re', building_mfa_system.FlowDict['Reuse of steel element'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of concrete element'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Reuse of concrete element'].Values)
    msf.ExcelSheetFill(results_file, 'C_el_re', building_mfa_system.FlowDict['Reuse of concrete element'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_reuseelements.xls')


def calc_reuse_steel(building_mfa_system, results_path, mylog):
    solve_mfa_reuse_steel(building_mfa_system, mylog)
    write_results_excel_reuse_steel(building_mfa_system, results_path, mylog)


def solve_mfa_reuse_steel(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to reuse of structural steel')
    print('Calculate reduced material demand due to reuse of structural steel')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Reuse of steel'].Values = \
        np.einsum('rft, rft -> rft', building_mfa_system.ParameterDict['par_steel_reuse'].Values,
                  building_mfa_system.FlowDict['Steel outflow'].Values)
    building_mfa_system.FlowDict['Steel inflow'].Values = building_mfa_system.FlowDict['Steel inflow'].Values + \
        building_mfa_system.FlowDict['Reuse of steel'].Values
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_reuse_steel(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE case (reuse of steel)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE case (reuse of steel)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of steel'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel'].Values)
    msf.ExcelSheetFill(results_file, 'S_reu', building_mfa_system.FlowDict['Reuse of steel'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_reusesteel.xls')


def calc_rec_cement(building_mfa_system, results_path, mylog):
    solve_mfa_rec_cement(building_mfa_system, mylog)
    write_results_excel_rec_cement(building_mfa_system, results_path, mylog)


def solve_mfa_rec_cement(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to cement recycling')
    print('Calculate reduced material demand due to cement recycling')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Cement recycling'].Values = \
        np.einsum('romt, rot -> rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_cement_recycling'].Values,
                            np.einsum('rot, rot ->rot',
                                      building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                                      (building_mfa_system.FlowDict['Concrete outflow'].Values +
                                       building_mfa_system.FlowDict['Concrete reuse'].Values))))
    building_mfa_system.FlowDict['Cement production'].Values = \
        building_mfa_system.FlowDict['Cement production'].Values + \
        building_mfa_system.FlowDict['Cement recycling'].Values
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values) - \
        np.einsum('rot, rot ->rt', building_mfa_system.ParameterDict['par_cement_recycling'].Values,
                  (building_mfa_system.FlowDict['Concrete outflow'].Values +
                   building_mfa_system.FlowDict['Concrete reuse'].Values))


def write_results_excel_rec_cement(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE case (cement recycling)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE case (cement recycling)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_rec', building_mfa_system.FlowDict['Cement recycling'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_cementrecycling.xls')


def ce_bundle_calculation(building_mfa_system, results_path, mylog):
    calc_bundle_lifestyle(building_mfa_system, results_path, mylog)
    calc_bundle_construction(building_mfa_system, results_path, mylog)
    calc_bundle_midway(building_mfa_system, results_path, mylog)


def calc_bundle_lifestyle(building_mfa_system, results_path, mylog):
    solve_mfa_bundle_lifestyle(building_mfa_system, mylog)
    write_results_excel_bundle_lifestyle(building_mfa_system, results_path, mylog)


def solve_mfa_bundle_lifestyle(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to CE bundle affecting lifestyle')
    print('Calculate reduced material demand due to CE bundle affecting lifestyle')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow_lifestyle'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock_lifestyle'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow_lifestyle'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_lifestyle'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_lifestyle'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_lifestyle'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_lifestyle'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_lifestyle'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_lifestyle'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values)


def write_results_excel_bundle_lifestyle(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE bundle (lifestyle)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE bundle (lifestyle)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_ lifestyle.xls')


def calc_bundle_construction(building_mfa_system, results_path, mylog):
    solve_mfa_bundle_construction(building_mfa_system, mylog)
    write_results_excel_bundle_construction(building_mfa_system, results_path, mylog)


def solve_mfa_bundle_construction(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to CE bundle affecting construction')
    print('Calculate reduced material demand due to CE bundle affecting construction')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_construction'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_construction'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_construction'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_construction'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_construction'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_construction'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Reuse of steel element'].Values = \
        np.einsum('rft, rft -> rft',
                  building_mfa_system.ParameterDict['par_steel_element_reuse'].Values,
                  building_mfa_system.FlowDict['Steel outflow'].Values)
    building_mfa_system.FlowDict['Steel inflow'].Values = building_mfa_system.FlowDict['Steel inflow'].Values + \
                                                          building_mfa_system.FlowDict['Reuse of steel element'].Values
    building_mfa_system.FlowDict['Steel outflow'].Values = building_mfa_system.FlowDict['Steel outflow'].Values - \
                                                           building_mfa_system.FlowDict['Reuse of steel element'].Values
    building_mfa_system.FlowDict['Reuse of steel'].Values = \
        np.einsum('rft, rft -> rft', building_mfa_system.ParameterDict['par_steel_reuse'].Values,
                  building_mfa_system.FlowDict['Steel outflow'].Values)
    building_mfa_system.FlowDict['Steel inflow'].Values = building_mfa_system.FlowDict['Steel inflow'].Values + \
                                                          building_mfa_system.FlowDict['Reuse of steel'].Values
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel element'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Reuse of concrete element'].Values = \
        np.einsum('rot, rot -> rot',
                  building_mfa_system.ParameterDict['par_concrete_element_reuse'].Values,
                  building_mfa_system.FlowDict['Concrete outflow'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        building_mfa_system.FlowDict['Concrete inflow'].Values + \
        building_mfa_system.FlowDict['Reuse of concrete element'].Values
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        building_mfa_system.FlowDict['Concrete outflow'].Values - \
        building_mfa_system.FlowDict['Reuse of concrete element'].Values
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Cement recycling'].Values = \
        np.einsum('romt, rot -> rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_cement_recycling'].Values,
                            np.einsum('rot, rot ->rot',
                                      building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                                      (building_mfa_system.FlowDict['Concrete outflow'].Values +
                                       building_mfa_system.FlowDict['Concrete reuse'].Values))))
    building_mfa_system.FlowDict['Cement production'].Values = \
        building_mfa_system.FlowDict['Cement production'].Values + \
        building_mfa_system.FlowDict['Cement recycling'].Values
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Reuse of concrete element'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values) - \
        np.einsum('rot, rot ->rt', building_mfa_system.ParameterDict['par_cement_recycling'].Values,
                  (building_mfa_system.FlowDict['Concrete outflow'].Values +
                   building_mfa_system.FlowDict['Concrete reuse'].Values))


def write_results_excel_bundle_construction(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE bundle (construction)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE bundle (construction)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of steel element'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel element'].Values)
    msf.ExcelSheetFill(results_file, 'S_el_re', building_mfa_system.FlowDict['Reuse of steel element'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of steel'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel'].Values)
    msf.ExcelSheetFill(results_file, 'S_reu', building_mfa_system.FlowDict['Reuse of steel'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of concrete element'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Reuse of concrete element'].Values)
    msf.ExcelSheetFill(results_file, 'C_el_re', building_mfa_system.FlowDict['Reuse of concrete element'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_rec', building_mfa_system.FlowDict['Cement recycling'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_construction.xls')


def calc_bundle_midway(building_mfa_system, results_path, mylog):
    solve_mfa_bundle_midway(building_mfa_system, mylog)
    write_results_excel_bundle_midway(building_mfa_system, results_path, mylog)


def solve_mfa_bundle_midway(building_mfa_system, mylog):
    mylog.info('Calculate reduced material demand due to "Midway" CE bundle')
    print('Calculate reduced material demand due to "Midway" CE bundle')
    building_mfa_system.FlowDict['Construction of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_inflow_midway'].Values
    building_mfa_system.StockDict['Building stock'].Values = \
        building_mfa_system.ParameterDict['par_building_stock_midway'].Values
    building_mfa_system.FlowDict['Demolition of buildings'].Values = \
        building_mfa_system.ParameterDict['par_building_outflow_midway'].Values
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rbaf, rbat->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_midway'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_midway'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rbaf, rbat ->rft',
                  building_mfa_system.ParameterDict['par_mi_steel_midway'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_midway'].Values,
                  building_mfa_system.FlowDict['Construction of buildings'].Values)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_midway'].Values,
                  building_mfa_system.StockDict['Building stock'].Values)
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        np.einsum('rbao, rbat ->rot',
                  building_mfa_system.ParameterDict['par_mi_concrete_midway'].Values,
                  building_mfa_system.FlowDict['Demolition of buildings'].Values)
    building_mfa_system.FlowDict['Reuse of steel element'].Values = \
        np.einsum('rft, rft -> rft',
                  building_mfa_system.ParameterDict['par_steel_element_reuse_midway'].Values,
                  building_mfa_system.FlowDict['Steel outflow'].Values)
    building_mfa_system.FlowDict['Steel inflow'].Values = building_mfa_system.FlowDict['Steel inflow'].Values + \
                                                          building_mfa_system.FlowDict['Reuse of steel element'].Values
    building_mfa_system.FlowDict['Steel outflow'].Values = building_mfa_system.FlowDict['Steel outflow'].Values - \
                                                           building_mfa_system.FlowDict['Reuse of steel element'].Values
    building_mfa_system.FlowDict['Reuse of steel'].Values = \
        np.einsum('rft, rft -> rft', building_mfa_system.ParameterDict['par_steel_reuse_midway'].Values,
                  building_mfa_system.FlowDict['Steel outflow'].Values)
    building_mfa_system.FlowDict['Steel inflow'].Values = building_mfa_system.FlowDict['Steel inflow'].Values + \
                                                          building_mfa_system.FlowDict['Reuse of steel'].Values
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rfst, rft ->rst',
                  building_mfa_system.ParameterDict['par_steel_process'].Values,
                  np.einsum('rft, rft->rft',
                            building_mfa_system.ParameterDict['par_finished_losses'].Values,
                            building_mfa_system.FlowDict['Steel inflow'].Values))
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst, rst ->rst',
                  building_mfa_system.ParameterDict['par_steel_recycling'].Values,
                  np.einsum('rst,  rst ->rst',
                            building_mfa_system.FlowDict['Steel production'].Values,
                            building_mfa_system.ParameterDict['par_steel_losses'].Values))
    building_mfa_system.FlowDict['Scrap other use'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel element'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel'].Values) - \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values) + \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values) - \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    building_mfa_system.FlowDict['Reuse of concrete element'].Values = \
        np.einsum('rot, rot -> rot',
                  building_mfa_system.ParameterDict['par_concrete_element_reuse_midway'].Values,
                  building_mfa_system.FlowDict['Concrete outflow'].Values)
    building_mfa_system.FlowDict['Concrete inflow'].Values = \
        building_mfa_system.FlowDict['Concrete inflow'].Values + \
        building_mfa_system.FlowDict['Reuse of concrete element'].Values
    building_mfa_system.FlowDict['Concrete outflow'].Values = \
        building_mfa_system.FlowDict['Concrete outflow'].Values - \
        building_mfa_system.FlowDict['Reuse of concrete element'].Values
    building_mfa_system.FlowDict['Cement production'].Values = \
        np.einsum('romt, rot ->rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Concrete reuse'].Values = \
        np.einsum('rot, rot ->rot',
                  building_mfa_system.ParameterDict['par_concrete_reuse'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                            building_mfa_system.FlowDict['Concrete inflow'].Values))
    building_mfa_system.FlowDict['Cement recycling'].Values = \
        np.einsum('romt, rot -> rmt',
                  building_mfa_system.ParameterDict['par_cement_process'].Values,
                  np.einsum('rot, rot ->rot',
                            building_mfa_system.ParameterDict['par_cement_recycling_midway'].Values,
                            np.einsum('rot, rot ->rot',
                                      building_mfa_system.ParameterDict['par_concrete_losses'].Values,
                                      (building_mfa_system.FlowDict['Concrete outflow'].Values +
                                       building_mfa_system.FlowDict['Concrete reuse'].Values))))
    building_mfa_system.FlowDict['Cement production'].Values = \
        building_mfa_system.FlowDict['Cement production'].Values + \
        building_mfa_system.FlowDict['Cement recycling'].Values
    building_mfa_system.FlowDict['Clinker production'].Values = \
        np.einsum('rmlt, rmt ->rlt',
                  building_mfa_system.ParameterDict['par_clinker_process'].Values,
                  np.einsum('rmt, rmt ->rmt',
                            building_mfa_system.ParameterDict['par_cement_losses'].Values,
                            building_mfa_system.FlowDict['Cement production'].Values))
    building_mfa_system.FlowDict['Concrete landfill'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete outflow'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Reuse of concrete element'].Values) + \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Concrete reuse'].Values) - \
        np.einsum('rot, rot ->rt', building_mfa_system.ParameterDict['par_cement_recycling_midway'].Values,
                  (building_mfa_system.FlowDict['Concrete outflow'].Values +
                   building_mfa_system.FlowDict['Concrete reuse'].Values))


def write_results_excel_bundle_midway(building_mfa_system, results_path, mylog):
    mylog.info('Export results to EXCEL for CE bundle (midway)')
    results_file = xlwt.Workbook()
    print('Export results to EXCEL for CE bundle (midway)')
    building_mfa_system.FlowDict['Steel inflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel inflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_in', building_mfa_system.FlowDict['Steel inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Steel stock in buildings'].Values = \
        np.einsum('rft->rt', building_mfa_system.StockDict['Steel stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'S_stock',
                       building_mfa_system.StockDict['Steel stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel outflow'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Steel outflow'].Values)
    msf.ExcelSheetFill(results_file, 'S_out', building_mfa_system.FlowDict['Steel outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of steel element'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel element'].Values)
    msf.ExcelSheetFill(results_file, 'S_el_re', building_mfa_system.FlowDict['Reuse of steel element'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of steel'].Values = \
        np.einsum('rft->rt', building_mfa_system.FlowDict['Reuse of steel'].Values)
    msf.ExcelSheetFill(results_file, 'S_reu', building_mfa_system.FlowDict['Reuse of steel'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Steel production'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Steel production'].Values)
    msf.ExcelSheetFill(results_file, 'S_prod', building_mfa_system.FlowDict['Steel production'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Scrap recycling'].Values = \
        np.einsum('rst->rt', building_mfa_system.FlowDict['Scrap recycling'].Values)
    msf.ExcelSheetFill(results_file, 'S_rec', building_mfa_system.FlowDict['Scrap recycling'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'S_oth', building_mfa_system.FlowDict['Scrap other use'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete inflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete inflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_in', building_mfa_system.FlowDict['Concrete inflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.StockDict['Concrete stock in buildings'].Values = \
        np.einsum('rot->rt', building_mfa_system.StockDict['Concrete stock in buildings'].Values)
    msf.ExcelSheetFill(results_file, 'C_stock', building_mfa_system.StockDict['Concrete stock in buildings'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete outflow'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete outflow'].Values)
    msf.ExcelSheetFill(results_file, 'C_out', building_mfa_system.FlowDict['Concrete outflow'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Reuse of concrete element'].Values = \
        np.einsum('rot->rt', building_mfa_system.FlowDict['Reuse of concrete element'].Values)
    msf.ExcelSheetFill(results_file, 'C_el_re', building_mfa_system.FlowDict['Reuse of concrete element'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CE_prod', building_mfa_system.FlowDict['Cement production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'CL_prod', building_mfa_system.FlowDict['Clinker production'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    building_mfa_system.FlowDict['Concrete reuse'].Values = np.einsum('rot->rt', building_mfa_system.FlowDict[
        'Concrete reuse'].Values)
    msf.ExcelSheetFill(results_file, 'C_reuse', building_mfa_system.FlowDict['Concrete reuse'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_rec', building_mfa_system.FlowDict['Cement recycling'].Values[:, 0, :],
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    msf.ExcelSheetFill(results_file, 'C_was', building_mfa_system.FlowDict['Concrete landfill'].Values,
                       rowlabels=building_mfa_system.IndexTable['Classification']['Region'].Items,
                       collabels=building_mfa_system.IndexTable['Classification']['Time'].Items)
    results_file.save(results_path + '/CE_midway.xls')


if __name__ == '__main__':
    main()
